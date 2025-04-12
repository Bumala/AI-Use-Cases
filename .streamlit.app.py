import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
uploaded_file = "Use Cases.xlsx"

# Load workbook using openpyxl
wb = load_workbook(uploaded_file, data_only=True)
sheet = wb.worksheets[1]  # Load the second sheet (0-based index)

# Extract the data and formatting
data = []
styles = []

# Iterate through rows and columns to extract values and styles
for row in sheet.iter_rows():
    row_data = []
    row_styles = []
    for cell in row:
        row_data.append(cell.value)
        cell_style = {
            "background_color": cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else None,
            "font_color": cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else "#000000",
            "bold": cell.font.bold,
            "italic": cell.font.italic,
        }
        row_styles.append(cell_style)
    data.append(row_data)
    styles.append(row_styles)

# Remove the first 3 rows and last 7 rows
data = data[3:-7]
styles = styles[3:-7]

# Convert data to a DataFrame
df = pd.DataFrame(data)

# Generate styled HTML table with rowspan for merged cells
def generate_html_table_with_rowspans(df, styles):
    html = "<table style='border-collapse: collapse; width: 100%;'>"
    for i, row in df.iterrows():
        html += "<tr>"
        for j, value in row.items():
            if j == 0:  # First column
                # Merge rows 1-5 for "Impact (What)"
                if i == 0:
                    style = styles[i][j]
                    background_color = f"background-color: #{style['background_color'][:6]};" if style['background_color'] else ""
                    font_color = f"color: #{style['font_color'][:6]};" if style['font_color'] else ""
                    font_weight = "font-weight: bold;" if style["bold"] else ""
                    font_style = "font-style: italic;" if style["italic"] else ""
                    cell_style = f"{background_color} {font_color} {font_weight} {font_style} padding: 5px; border: 1px solid #ddd;"
                    html += f"<td rowspan='5' style='{cell_style}'>Impact (What)</td>"
                elif i < 5:
                    continue  # Skip rows 2 to 5

                # Merge rows 6-10 for "Technology (How)"
                elif i == 5:
                    style = styles[i][j]
                    background_color = f"background-color: #{style['background_color'][:6]};" if style['background_color'] else ""
                    font_color = f"color: #{style['font_color'][:6]};" if style['font_color'] else ""
                    font_weight = "font-weight: bold;" if style["bold"] else ""
                    font_style = "font-style: italic;" if style["italic"] else ""
                    cell_style = f"{background_color} {font_color} {font_weight} {font_style} padding: 5px; border: 1px solid #ddd;"
                    html += f"<td rowspan='5' style='{cell_style}'>Technology (How)</td>"
                elif 5 < i < 10:
                    continue  # Skip rows 7 to 10

                # Merge the last two rows for "Place (Where)"
                elif i == len(df) - 2:  # Second to last row
                    style = styles[i][j]
                    background_color = f"background-color: #{style['background_color'][:6]};" if style['background_color'] else ""
                    font_color = f"color: #{style['font_color'][:6]};" if style['font_color'] else ""
                    font_weight = "font-weight: bold;" if style["bold"] else ""
                    font_style = "font-style: italic;" if style["italic"] else ""
                    cell_style = f"{background_color} {font_color} {font_weight} {font_style} padding: 5px; border: 1px solid #ddd;"
                    html += f"<td rowspan='2' style='{cell_style}'>Place (Where)</td>"
                elif i == len(df) - 1:  # Last row
                    continue  # Skip the last row in the first column

                # Render remaining cells in the first column
                else:
                    style = styles[i][j]
                    background_color = f"background-color: #{style['background_color'][:6]};" if style['background_color'] else ""
                    font_color = f"color: #{style['font_color'][:6]};" if style['font_color'] else ""
                    font_weight = "font-weight: bold;" if style["bold"] else ""
                    font_style = "font-style: italic;" if style["italic"] else ""
                    cell_style = f"{background_color} {font_color} {font_weight} {font_style} padding: 5px; border: 1px solid #ddd;"
                    html += f"<td style='{cell_style}'>{value if value is not None else ''}</td>"
            else:  # Other columns
                style = styles[i][j]
                background_color = f"background-color: #{style['background_color'][:6]};" if style['background_color'] else ""
                font_color = f"color: #{style['font_color'][:6]};" if style['font_color'] else ""
                font_weight = "font-weight: bold;" if style["bold"] else ""
                font_style = "font-style: italic;" if style["italic"] else ""
                cell_style = f"{background_color} {font_color} {font_weight} {font_style} padding: 5px; border: 1px solid #ddd;"
                html += f"<td style='{cell_style}'>{value if value is not None else ''}</td>"
        html += "</tr>"
    html += "</table>"
    return html

# Streamlit app
st.set_page_config(layout="wide", page_title="Morphological Box Viewer")
st.title("Morphological Box Viewer")

# Display the styled morphological box
st.markdown("### Morphological Box")
html_table = generate_html_table_with_rowspans(df, styles)
st.markdown(html_table, unsafe_allow_html=True)

st.markdown("---")
st.info("This morphological box is displayed with the formatting extracted from the Excel file.")
